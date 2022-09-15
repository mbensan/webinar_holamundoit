import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

# ajuste para trabajar con OpenPyXL
import collections
collections.Callable = collections.abc.Callable

URL = 'https://chile.as.com/resultados/futbol/chile/calendario/'


# para crear mi archivo Excel
wb = Workbook()
ws = wb.active


def writeJornada(titulo, fecha, filas):
    # 0. Elijo un punto de partida
    inicio = ws.max_row + 2

    # 1. Escribo el titulo en negritas
    ws['B' + str(inicio)] = titulo
    ws['B' + str(inicio)].font = Font(bold=True)

    # 2. Escribo la fecha en Negritas
    ws['C' + str(inicio)] = fecha
    ws['C' + str(inicio)].font = Font(bold=True)

    # 3. Escribo los encabezados
    ws['B' + str(inicio + 1)] = 'Local'
    ws['B' + str(inicio + 1)].font = Font(bold=True)

    ws['C' + str(inicio + 1)] = 'Resultado'
    ws['C' + str(inicio + 1)].font = Font(bold=True)

    ws['D' + str(inicio + 1)] = 'Visita'
    ws['D' + str(inicio + 1)].font = Font(bold=True)

    # 4. Escribo los partidos
    num_fila = inicio + 2
    #import pdb; pdb.set_trace()
    for fila in filas:

        ws['B' + str(num_fila)] = fila['equipo-local']
        ws['C' + str(num_fila)] = fila['resultado']
        ws['D' + str(num_fila)] = fila['equipo-visitante']
        num_fila += 1


def parseJornada(jornada):
    # 1. Extraemos el titulo y la fecha de la jornada
    titulo = jornada.select('.tit-modulo a')[0].text
    fecha = jornada.select('.tit-modulo span')[0].text

    # 2. Eliminamos espacios en blanco de la fecha
    fecha = fecha.strip()

    # 3. Extraemos los tr's (los partidos) de cada jornada
    trs = jornada.select('.tabla-datos tbody tr')

    # 4. Vamos a crear las filas para esribir en el Excel
    filas = []
    for tr in trs:
        new_fila = {
            'equipo-local': tr.select('.col-equipo-local')[0].text.strip(),
            'resultado': tr.select('.col-resultado')[0].text.strip(),
            'equipo-visitante': tr.select('.col-equipo-visitante')[0].text.strip()
        }
        filas.append(new_fila)

    writeJornada(titulo, fecha, filas)


def main():
    page = requests.get(URL)
    # 1. parseamos el contenido del HTML
    soup = BeautifulSoup(page.content, 'html.parser')
    # 2. creamos una lista de python, con todas las 30 jornadas
    jornadas = []
    for i in range(1, 31):
        nueva_jornada = soup.find(id='jornada-' + str(i))
        jornadas.append(nueva_jornada)
    
    for jornada in jornadas:
        parseJornada(jornada)

    ahora = datetime.now()
    fecha_archivo = ahora.strftime('%Y_%m_%d_%H_%M_%S')
    wb.save('partidos_' + fecha_archivo + '.xlsx')


main()
